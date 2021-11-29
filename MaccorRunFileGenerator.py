from tkinter import Tk  # from tkinter import Tk for Python 3.x
import tkinter as tk
from tkinter import ttk
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
import random
import string
from itertools import islice
import pandas as pd


class MaccorRunFileGenerator():
    def __init__(self):
        self.file = ''
        self.getFileName()
        self.wb = 'file_holder'
        self.wb_loaded = False
        self.openWorbookFile()
        self.root = tk.Tk()
        self.choiceVar = tk.StringVar()
        self.sheet = ''

        self.channelStart = tk.StringVar()
        self.procedureName = tk.StringVar()
        self.rowStart = tk.StringVar()
        self.cellCount = tk.StringVar()
        # self.resetVars()

        self.channelsList = []
        self.procedureNameList = []
        self.rowStartList = []
        self.cellCountList = []
        self.commonProcedureNames = []
        # self.initCommonProcedureNames()

        self.inputFrame = tk.Frame(self.root)
        self.inputFrameInstance = 0

        self.currentSheet = ''

    def getFileName(self):
        Tk().withdraw()  # we don't want a full GUI, so keep the root window from appearing
        filename = askopenfilename()  # show an "Open" dialog box and return the path to the selected file
        print(filename)
        self.file = filename
        # self.root.quit()

    def openWorbookFile(self):

        """
        try:
            self.wb = load_workbook(filename=self.file)
            self.wb_loaded = True
        except:
            print('Could not load workbook')
        """
        try:
            self.wb = pd.read_excel(self.file, sheet_name=None, usecols='A:FZ')
            #self.wb_path = self.file
            self.wb_loaded = True
            #print(self.wb['EV full cells'].head())
        except:
            print('Could not load workbook')


    def selectWorkbookSheet(self):
        if self.wb_loaded:
            #print(self.wb.sheetnames)
            print(self.wb.keys())
            # root = tk.Tk()
            """
            self.checkbar = Checkbar(parent=self.root, picks=self.wb.sheetnames)
            self.checkbar.pack()
            self.sheet = self.checkbar.state()
            print('self.sheet is: '+str(self.sheet))
            """

            #choices = self.wb.sheetnames
            choices = list(self.wb.keys())
            #self.choiceVar.set(choices[0])

            # om = tk.OptionMenu(self, self.choiceVar, *choices)
            self.cb = ttk.Combobox(self.root, textvariable=self.choiceVar, values=choices, width=50)
            # om.pack()
            self.cb.pack()
            self.selectSheetButton = ttk.Button(self.root, text='Select Sheet', command=self.selectSheetButton)
            self.selectSheetButton.pack()
            # return

    def selectSheetButton(self):
        if self.wb_loaded:
            self.sheet = self.cb.get()
            print('cb contents is: ')
            print(self.cb.get())
            self.currentSheet = self.wb[self.sheet]
            self.genPage()

    def changeSheet(self):
        self.currentSheet = self.wb[self.sheet]

    def genPage(self):
        if self.inputFrameInstance < 1:
            # self.inputFrame.pack()
            print('generating page, if')
            print('current sheet - '+ self.sheet)
            """
            self.channelEntry = tk.Entry(master=self.inputFrame, textvariable=self.channelStart)
            self.channelEntry.pack()
            self.procedureNameEntry= tk.Entry(master=self.inputFrame, textvariable=self.procedureName)
            self.procedureNameEntry.pack()
            self.rowStartEntry = tk.Entry(master=self.inputFrame, textvariable=self.rowStart)
            self.rowStartEntry.pack()
            self.cellCount = tk.Entry(master=self.inputFrame, textvariable=self.cellCount)
            self.cellCountEntry.pack()
            
            self.cb = ttk.Combobox(self.root, textvariable=self.choiceVar, values=choices)
            """
            # self.valuesUpdate()
            self.channelCb = ttk.Combobox(master=self.inputFrame, textvariable=self.channelStart,
                                          values=self.genchannelsList(),
                                          postcommand=self.valuesUpdate('channel'),
                                          width=50)
            self.channelCb.pack()
            self.procedureNameCb = ttk.Combobox(master=self.inputFrame, textvariable=self.procedureName,
                                                values=self.genprocedureNameList(),
                                                postcommand=self.valuesUpdate('procedure'),
                                                width=50)
            self.procedureNameCb.pack()
            self.rowStartCb = ttk.Combobox(master=self.inputFrame, textvariable=self.rowStart,
                                           values=self.genrowStartList(),
                                           postcommand=self.valuesUpdate('row'),
                                           width=50)  # values = self.wb[self.sheet])
            self.rowStartCb.pack()
            self.cellCountCb = ttk.Combobox(master=self.inputFrame, textvariable=self.cellCount,
                                            values=self.gencellCountList(),
                                            postcommand=self.valuesUpdate('cell'),
                                            width=50)
            self.cellCountCb.pack()

            self.createTxtButton = ttk.Button(self.inputFrame, text='Generate File', command=self.createTxt)
            self.createTxtButton.pack()

            self.inputFrame.pack()
            self.inputFrameInstance = self.inputFrameInstance + 1
        else:
            print('generating page, else')
            print('current sheet - ' + self.sheet)
            for widget in self.inputFrame.winfo_children():
                widget.destroy()
            self.channelCb = ttk.Combobox(master=self.inputFrame, textvariable=self.channelStart,
                                          values=self.genchannelsList(),
                                          postcommand=self.valuesUpdate('channel'),
                                          width=50)
            self.channelCb.pack()
            self.procedureNameCb = ttk.Combobox(master=self.inputFrame, textvariable=self.procedureName,
                                                values=self.genprocedureNameList(),
                                                postcommand=self.valuesUpdate('procedure'),
                                                width=50)
            self.procedureNameCb.pack()
            self.rowStartCb = ttk.Combobox(master=self.inputFrame, textvariable=self.rowStart,
                                           values=self.genrowStartList(),
                                           postcommand=self.valuesUpdate('row'),
                                           width=50)  # values = self.wb[self.sheet])
            self.rowStartCb.pack()
            self.cellCountCb = ttk.Combobox(master=self.inputFrame, textvariable=self.cellCount,
                                            values=self.gencellCountList(),
                                            postcommand=self.valuesUpdate('cell'),
                                            width=50)
            self.cellCountCb.pack()
            self.createTxtButton = ttk.Button(self.inputFrame, text='Generate File', command=self.createTxt)
            self.createTxtButton.pack()

            self.inputFrame.pack()

    def resetVars(self):
        print('resetVars call')
        self.channelStart.set('First Channel Here - Will increment by 1')
        self.procedureName.set('Procedure Name Here')
        self.rowStart.set('Row Start from Coin Cell Database Here')
        self.cellCount.set('Number of Cells Here')

    def valuesUpdate(self, cb=None):
        print('Values Update Call')
        print(cb)
        if cb == None:
            self.channelsList = []
            self.procedureNameList = []
            self.rowStartList = []
            self.cellCountList = []

            self.setFirstValues()
            self.channelsList.extend(list(range(1, 64 + 1)))
            self.cellCountList.extend(list(range(1, 64 + 1)))
            self.procedureNameList.append(self.initCommonProcedureNames())
            self.rowStartList.append(self.getLastTenCellNames())

        elif cb == 'channel':
            self.channelsList = []
            self.channelsList.append('First Channel Here - Will increment by 1')
            self.channelsList.extend(list(range(1, 64 + 1)))

        elif cb == 'procedure':
            self.procedureNameList = []
            self.procedureNameList.append('Procedure Name Here')
            self.procedureNameList.append(self.initCommonProcedureNames())

        elif cb == 'row':
            self.rowStartList = []
            self.rowStartList.append('Row Start from Coin Cell Database Here')
            self.rowStartList.append(self.getLastTenCellNames())

        elif cb == 'cell':
            self.cellCountList = []
            self.cellCountList.append('Number of Cells Here')
            self.cellCountList.extend(list(range(1, 64 + 1)))
        self.printLists()

    def printLists(self):
        print('print list call')
        print(self.channelsList)
        print(self.procedureNameList)
        print(self.rowStartList)
        print(self.cellCountList)

    def genchannelsList(self):
        channelsList = []
        channelsList.append('First Channel Here - Will increment by 1')
        channelsList.extend(list(range(1, 64 + 1)))
        return channelsList

    def genprocedureNameList(self):
        procedureNameList = []
        procedureNameList.append('Procedure Name Here')
        procedureNameList.extend(self.initCommonProcedureNames())
        return procedureNameList

    def genrowStartList(self):
        rowStartList = []
        rowStartList.append('Row Start from Coin Cell Database Here')
        rowStartList.extend(self.getLastTenCellNames())
        return rowStartList

    def gencellCountList(self):
        cellCountList = []
        cellCountList.append('Number of Cells Here')
        cellCountList.extend(list(range(1, 64 + 1)))
        return cellCountList

    def setFirstValues(self):
        self.channelsList.append('First Channel Here - Will increment by 1')
        self.procedureNameList.append('Procedure Name Here')
        self.rowStartList.append('Row Start from Coin Cell Database Here')
        self.cellCountList.append('Number of Cells Here')

        print('setFirstValuesCall')
        print(self.channelsList)
        print(self.procedureNameList)
        print(self.rowStartList)
        print(self.cellCountList)

    def initCommonProcedureNames(self):
        commonProcedures = ['DOE_EV Half_0,1C form and cycling.000']
        return commonProcedures

    def getLastTenCellNames(self):
        print('getLast10 Call')
        lastTenCellNames = []
        #data = pd.read_excel(self.file, sheet_name=self.cb.get(), index_col=0)
        data = self.wb[self.cb.get()]
        last = data.tail(30)
        lastCell = last.iloc[:, 0].tolist()
        print('lastTenCellNames: ')
        print(list(lastCell))
        print(type(lastCell))
        lastTenCellNames=lastCell

        #letters = string.ascii_lowercase
        #lastTenCellNames.append(''.join(random.choice(letters) for i in range(10)))
        # last_empty_row = len(list(self.wb[self.sheet].rows))

        return lastTenCellNames

    def createTxt(self):
        print('createTxt call')
        channelStart = self.channelCb.get()
        procedureName = self.procedureNameCb.get()
        rowStart = self.rowStartCb.get()
        cellCount = self.cellCountCb.get()

        if self.cbCheck():
            self.writeTxtFile(channelStart=channelStart, procedureName=procedureName,
                          rowStart=rowStart, cellCount=cellCount)

    def writeTxtFile(self, channelStart = 1, procedureName = 'Procedure Name',
                     rowStart = 'E0559', cellCount = 1):

        """
        data = self.currentSheet.values
        cols = next(data)[1:]
        data = list(data)
        idx = [r[0] for r in data]
        data = (islice(r, 1, None) for r in data)
        df = pd.DataFrame(data, index=idx, columns=cols)
        print(df.head())
        """
        print('current sheet is: ')
        print(self.cb.get())
        #print(self.wb)
        #data = pd.read_excel(self.file, sheet_name=self.cb.get(), )#index_col=0)
        data = self.wb[self.cb.get()]
        #data.reset_index()
        print(data.head(10))
        print('current index is: ')
        print(data[data.iloc[:, 0] == rowStart].index[0])

        first_index = data[data.iloc[:, 0] == rowStart].index[0]

        filename = self.txtFileNameGenerator()
        writefile = open(filename, 'w')
        final_index = int(first_index)+int(cellCount)

        if self.checkIndexPlusCells():
            channel = channelStart
            for current_index in range(int(first_index), final_index):
                print('write file loop: '+str(current_index))
                write_string = ''
                tab = '\t'
                newline = '\n'

                write_string = write_string+channel+tab
                channel = str(int(channel)+1)

                cellname = data[data.iloc[:, 0].index == current_index].iloc[:, 0].item()
                print(cellname)
                print('cellname is: '+cellname)
                write_string = write_string+cellname+tab

                procedure = self.procedureNameCb.get()
                write_string = write_string+procedure+tab

                cap = 'capacity [Ah]'
                mass = 'Active mass [g]'

                c_rate = str(data[data.loc[:, cap].index == current_index].loc[:, cap].item())
                print('c_rate is: ')
                print(c_rate)
                print('c_rate is: '+c_rate)
                write_string = write_string+c_rate+tab

                weight = str(data[data.loc[:, mass].index == current_index].loc[:, mass].item())
                print('weight is: '+ weight)
                write_string = write_string+weight+tab

                comment = cellname
                write_string = str(write_string+comment+newline)

                print('write_string is: '+str(write_string))

                writefile.write(write_string)

        writefile.close()



    def checkIndexPlusCells(self):

        return True

    def txtFileNameGenerator(self):
        file_name = 'temp.txt'

        return file_name

    def cbCheck(self):
        channelStart = self.channelStart.get()
        procedureName = self.procedureName.get()
        rowStart = self.rowStart.get()
        cellCount = self.cellCount.get()

        if True:  # Good condition here

            return True
        else:
            return False




class Checkbar(tk.Frame):
    def __init__(self, parent=None, picks=[], side=tk.LEFT, anchor=tk.W):
        tk.Frame.__init__(self, parent)
        self.vars = []
        for pick in picks:
            var = tk.IntVar()
            chk = tk.Checkbutton(self, text=pick, variable=var)
            chk.pack()
            self.vars.append(var)

    def state(self):
        return map((lambda var: var.get()), self.vars)


"""
test= file( "boats.tab", "w" )
test.write( "\t".join( Boat.csvHeading ) )
test.write( "\n" )
for d in db:
    test.write( "\t".join( map( str, d.csvRow() ) ) )
    test.write( "\n" )
test.close()
"""

if __name__ == "__main__":
    generator = MaccorRunFileGenerator()
    generator.selectWorkbookSheet()
    generator.root.mainloop()
    generator.root.quit()
